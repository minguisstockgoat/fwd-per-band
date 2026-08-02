#!/usr/bin/env python3
"""엑셀(코200, 닥150 12mf eps.xlsx) -> raw/eps_history.json 시드 변환.

엑셀 구조 (Sheet1):
  row 8  : Code   | A005930 | A000660 | ...
  row 9  : Name   | 삼성전자 | SK하이닉스 | ...
  row 14 : 헤더   | D A T E | EPS(Fwd.12M, 지배) ...
  row 15~: 날짜   | 값 ...
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path(
    r"C:\Users\eogks\OneDrive\바탕 화면\주식\다올프랍\코200, 닥150 12mf eps.xlsx"
)


def parse(xlsx: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["Sheet1"]

    header = list(ws.iter_rows(min_row=8, max_row=9, values_only=True))
    codes_row, names_row = header[0], header[1]

    cols: list[tuple[int, str, str]] = []  # (col_index, code, name)
    seen: set[str] = set()
    for idx, raw_code in enumerate(codes_row):
        if idx == 0 or not raw_code:
            continue
        code = str(raw_code).strip()
        if code.startswith("A"):
            code = code[1:]
        if not (len(code) == 6 and code.isdigit()):
            continue
        if code in seen:  # 엑셀에 같은 종목이 두 번 들어간 열이 있다
            continue
        seen.add(code)
        cols.append((idx, code, str(names_row[idx]).strip()))

    dates: list[str] = []
    series: dict[str, list[float | None]] = {c: [] for _, c, _ in cols}

    for row in ws.iter_rows(min_row=15, values_only=True):
        day = row[0]
        if day is None:
            continue
        dates.append(day.strftime("%Y%m%d"))
        for idx, code, _ in cols:
            val = row[idx] if idx < len(row) else None
            if isinstance(val, (int, float)) and val not in (0,):
                series[code].append(round(float(val), 4))
            else:
                series[code].append(None)

    return {
        "source": xlsx.name,
        "dates": dates,
        "names": {code: name for _, code, name in cols},
        "eps": series,
    }


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--out", type=Path, default=ROOT / "raw" / "eps_history.json")
    args = ap.parse_args()

    data = parse(args.xlsx)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
    print(
        f"stocks={len(data['eps'])} dates={len(data['dates'])} "
        f"({data['dates'][0]}~{data['dates'][-1]}) -> {args.out}"
    )


if __name__ == "__main__":
    main()
